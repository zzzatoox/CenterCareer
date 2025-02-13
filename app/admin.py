from django.contrib import admin
from .models import (
    City,
    Company,
    Vacancy,
    Filter,
    FilterValue,
    VacancyFilter,
    Status,
    Category,
    Location,
    Event,
)

from django_ckeditor_5.widgets import CKEditor5Widget
from django import forms

from openpyxl import load_workbook
from django.http import HttpResponseRedirect
from django.contrib import messages
from django.urls import path
from django.shortcuts import render, redirect

from django.db import transaction


class VacancyAdminForm(forms.ModelForm):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["description"].required = False

    description = forms.CharField(
        label="Описание",
        widget=CKEditor5Widget(
            attrs={"class": "django_ckeditor_5"}, config_name="default"
        ),
    )

    class Meta:
        model = Vacancy
        fields = "__all__"


class EventAdminForm(forms.ModelForm):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["description"].required = False

    description = forms.CharField(
        label="Описание",
        widget=CKEditor5Widget(
            attrs={"class": "django_ckeditor_5"}, config_name="default"
        ),
    )

    class Meta:
        model = Vacancy
        fields = "__all__"


@admin.register(City)
class CityAdmin(admin.ModelAdmin):
    list_display = ("title", "region")
    search_fields = ("title", "region")
    ordering = ("title",)


@admin.register(Company)
class CompanyAdmin(admin.ModelAdmin):
    list_display = ("title", "address", "city")
    list_filter = ("city",)
    search_fields = ("title", "address")
    ordering = ("title",)


class VacancyFilterInline(admin.TabularInline):
    model = VacancyFilter
    extra = 1


class CsvImportForm(forms.Form):
    excel_file = forms.FileField()


@admin.register(Vacancy)
class VacancyAdmin(admin.ModelAdmin):
    list_display = ("title", "company", "min_salary", "max_salary", "is_relevant")
    form = VacancyAdminForm
    list_filter = ("company", "is_relevant")
    search_fields = ("title", "description")
    ordering = ("title",)
    inlines = [VacancyFilterInline]
    change_list_template = "admin/vacancy_change_list.html"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("import-excel/", self.import_excel, name="import_excel"),
            path("confirm-import/", self.confirm_import, name="confirm_import"),
        ]
        return custom_urls + urls

    from django.db import transaction

    def confirm_import(self, request):
        if request.method == "POST" and request.POST.get("confirm") == "1":
            # Получаем данные из сессии
            preview_data = request.session.get("preview_data", [])

            if not preview_data:
                self.message_user(
                    request, "Нет данных для импорта.", level=messages.ERROR
                )
                return HttpResponseRedirect("..")

            errors = []
            try:
                with transaction.atomic():  # Начало транзакции
                    for row in preview_data:
                        try:
                            # Поиск компании
                            company = Company.objects.get(title=row["company"])

                            # Создание вакансии
                            vacancy = Vacancy.objects.create(
                                title=row["title"],
                                description=row["description"],
                                min_salary=row["min_salary"],
                                max_salary=row["max_salary"],
                                is_relevant=row["is_relevant"],
                                company=company,
                            )

                            # Добавление фильтров
                            if row["experience"]:
                                try:
                                    filter_value = FilterValue.objects.get(
                                        value=row["experience"],
                                        filter__title="Опыт работы",
                                    )
                                    VacancyFilter.objects.create(
                                        vacancy=vacancy, filter_value=filter_value
                                    )
                                except FilterValue.DoesNotExist:
                                    errors.append(
                                        f"Строка {preview_data.index(row) + 1}: Значение фильтра 'Опыт работы' '{row['experience']}' не найдено в базе данных."
                                    )

                            if row["employment_type"]:
                                try:
                                    filter_value = FilterValue.objects.get(
                                        value=row["employment_type"],
                                        filter__title="Тип занятости",
                                    )
                                    VacancyFilter.objects.create(
                                        vacancy=vacancy, filter_value=filter_value
                                    )
                                except FilterValue.DoesNotExist:
                                    errors.append(
                                        f"Строка {preview_data.index(row) + 1}: Значение фильтра 'Тип занятости' '{row['employment_type']}' не найдено в базе данных."
                                    )

                            if row["work_schedule"]:
                                try:
                                    filter_value = FilterValue.objects.get(
                                        value=row["work_schedule"],
                                        filter__title="Режим работы",
                                    )
                                    VacancyFilter.objects.create(
                                        vacancy=vacancy, filter_value=filter_value
                                    )
                                except FilterValue.DoesNotExist:
                                    errors.append(
                                        f"Строка {preview_data.index(row) + 1}: Значение фильтра 'Режим работы' '{row['work_schedule']}' не найдено в базе данных."
                                    )

                            if row["disability_friendly"]:
                                try:
                                    filter_value = FilterValue.objects.get(
                                        value=row["disability_friendly"],
                                        filter__title="Подходит для инвалидов",
                                    )
                                    VacancyFilter.objects.create(
                                        vacancy=vacancy, filter_value=filter_value
                                    )
                                except FilterValue.DoesNotExist:
                                    errors.append(
                                        f"Строка {preview_data.index(row) + 1}: Значение фильтра 'Подходит для инвалидов' '{row['disability_friendly']}' не найдено в базе данных."
                                    )

                        except Company.DoesNotExist:
                            errors.append(
                                f"Строка {preview_data.index(row) + 1}: Компания '{row['company']}' не найдена в базе данных. Пожалуйста, сначала добавьте компанию."
                            )
                        except Exception as e:
                            if "city_id" in str(e):
                                errors.append(
                                    f"Строка {preview_data.index(row) + 1}: У вас указана новая компания '{row['company']}', которой не существует в базе данных. Необходимо предварительно занести её в базу данных, чтобы загрузить для неё вакансии."
                                )
                            else:
                                errors.append(
                                    f"Строка {preview_data.index(row) + 1}: Ошибка при создании вакансии: {e}"
                                )

                    if errors:
                        raise Exception(
                            "Ошибки при импорте данных."
                        )  # Отмена транзакции

                    self.message_user(
                        request, "Данные успешно импортированы.", level=messages.SUCCESS
                    )

            except Exception as e:
                for error in errors:
                    self.message_user(request, error, level=messages.ERROR)
                self.message_user(
                    request, f"Ошибка при импорте данных: {e}", level=messages.ERROR
                )

            # Очищаем данные из сессии
            if "preview_data" in request.session:
                del request.session["preview_data"]

            return HttpResponseRedirect("..")

        return HttpResponseRedirect("..")

    def import_excel(self, request):
        if request.method == "POST":
            excel_file = request.FILES["excel_file"]
            wb = load_workbook(excel_file)
            ws = wb.active

            # Ожидаемые заголовки столбцов на русском
            expected_columns = [
                "Название",
                "Описание",
                "Минимальная зарплата",
                "Максимальная зарплата",
                "Актуальность",
                "Компания",
                "Опыт работы",
                "Тип занятости",
                "Режим работы",
                "Подходит для инвалидов",
            ]

            # Получаем заголовки из первой строки файла
            header_row = [cell.value for cell in ws[1]]

            # Проверяем, что заголовки соответствуют ожидаемым
            if header_row != expected_columns:
                self.message_user(
                    request,
                    "Неправильная структура файла. Проверьте столбцы.",
                    level=messages.ERROR,
                )
                return HttpResponseRedirect("..")

            errors = []
            preview_data = []  # Данные для предпросмотра
            for row_idx, row in enumerate(
                ws.iter_rows(min_row=2, values_only=True), start=2
            ):
                try:
                    # Сопоставляем русские заголовки с переменными
                    data = dict(zip(expected_columns, row))

                    # Извлекаем данные
                    title = data["Название"]
                    description = data["Описание"]
                    min_salary = data["Минимальная зарплата"]
                    max_salary = data["Максимальная зарплата"]
                    is_relevant = data["Актуальность"]
                    company_name = data["Компания"]
                    experience = data["Опыт работы"]
                    employment_type = data["Тип занятости"]
                    work_schedule = data["Режим работы"]
                    disability_friendly = data["Подходит для инвалидов"]

                    # Проверка обязательных полей
                    if not title:
                        raise ValueError(
                            f"Строка {row_idx}: Поле 'Название' обязательно для заполнения."
                        )
                    if not company_name:
                        raise ValueError(
                            f"Строка {row_idx}: Поле 'Компания' обязательно для заполнения."
                        )

                    # Проверка числовых полей
                    if min_salary is not None and not isinstance(
                        min_salary, (int, float)
                    ):
                        raise ValueError(
                            f"Строка {row_idx}: Поле 'Минимальная зарплата' должно быть числом."
                        )
                    if max_salary is not None and not isinstance(
                        max_salary, (int, float)
                    ):
                        raise ValueError(
                            f"Строка {row_idx}: Поле 'Максимальная зарплата' должно быть числом."
                        )

                    # Проверка булевого поля
                    if is_relevant not in (0, 1):
                        raise ValueError(
                            f"Строка {row_idx}: Поле 'Актуальность' должно быть 0 или 1."
                        )

                    # Поиск компании (в режиме предпросмотра не сохраняем)
                    try:
                        company = Company.objects.get(title=company_name)
                    except Company.DoesNotExist:
                        raise ValueError(
                            f"Строка {row_idx}: Компания '{company_name}' не найдена в базе данных. Пожалуйста, сначала добавьте компанию."
                        )

                    # Проверка значений фильтров
                    if experience:
                        try:
                            FilterValue.objects.get(
                                value=experience, filter__title="Опыт работы"
                            )
                        except FilterValue.DoesNotExist:
                            raise ValueError(
                                f"Строка {row_idx}: Значение фильтра 'Опыт работы' '{experience}' не найдено в базе данных."
                            )

                    if employment_type:
                        try:
                            FilterValue.objects.get(
                                value=employment_type, filter__title="Тип занятости"
                            )
                        except FilterValue.DoesNotExist:
                            raise ValueError(
                                f"Строка {row_idx}: Значение фильтра 'Тип занятости' '{employment_type}' не найдено в базе данных."
                            )

                    if work_schedule:
                        try:
                            FilterValue.objects.get(
                                value=work_schedule, filter__title="Режим работы"
                            )
                        except FilterValue.DoesNotExist:
                            raise ValueError(
                                f"Строка {row_idx}: Значение фильтра 'Режим работы' '{work_schedule}' не найдено в базе данных."
                            )

                    if disability_friendly:
                        try:
                            FilterValue.objects.get(
                                value=disability_friendly,
                                filter__title="Подходит для инвалидов",
                            )
                        except FilterValue.DoesNotExist:
                            raise ValueError(
                                f"Строка {row_idx}: Значение фильтра 'Подходит для инвалидов' '{disability_friendly}' не найдено в базе данных."
                            )

                    # Сбор данных для предпросмотра
                    preview_data.append(
                        {
                            "title": title,
                            "description": description,
                            "min_salary": min_salary,
                            "max_salary": max_salary,
                            "is_relevant": is_relevant,
                            "company": company_name,
                            "experience": experience,
                            "employment_type": employment_type,
                            "work_schedule": work_schedule,
                            "disability_friendly": disability_friendly,
                        }
                    )

                except ValueError as e:
                    errors.append(str(e))
                except FilterValue.DoesNotExist as e:
                    errors.append(
                        f"Строка {row_idx}: Некорректное значение фильтра: {e}"
                    )
                except Exception as e:
                    errors.append(f"Строка {row_idx}: Неизвестная ошибка: {e}")

            if errors:
                # Если есть ошибки, передаем их в шаблон
                return render(
                    request,
                    "admin/excel_import_form.html",
                    {"form": CsvImportForm(), "errors": errors},
                )

            # Сохраняем данные для предпросмотра в сессии
            request.session["preview_data"] = preview_data

            # Если ошибок нет, отображаем предпросмотр
            return render(
                request,
                "admin/excel_import_preview.html",
                {"preview_data": preview_data},
            )

        # Если метод GET, отображаем форму загрузки
        form = CsvImportForm()
        return render(
            request, "admin/excel_import_form.html", {"form": form, "errors": []}
        )


@admin.register(Filter)
class FilterAdmin(admin.ModelAdmin):
    list_display = ("title",)
    search_fields = ("title",)
    ordering = ("title",)


@admin.register(FilterValue)
class FilterValueAdmin(admin.ModelAdmin):
    list_display = ("filter", "value")
    list_filter = ("filter",)
    search_fields = ("value",)
    ordering = ("filter__title", "value")


@admin.register(VacancyFilter)
class VacancyFilterAdmin(admin.ModelAdmin):
    list_display = ("vacancy", "filter_value")
    list_filter = ("vacancy", "filter_value__filter")
    search_fields = ("vacancy__title", "filter_value__value")


@admin.register(Status)
class StatusAdmin(admin.ModelAdmin):
    list_display = ("title",)
    search_fields = ("title",)
    ordering = ("title",)


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ("title",)
    search_fields = ("title",)
    ordering = ("title",)


@admin.register(Location)
class LocationAdmin(admin.ModelAdmin):
    list_display = ("location",)
    search_fields = ("location",)
    ordering = ("location",)


@admin.register(Event)
class EventAdmin(admin.ModelAdmin):
    list_display = (
        "title",
        "event_date",
        "end_date",
        "event_month",
        "start_time",
        "end_time",
        "location",
        "category",
        "status",
        "company",
    )
    list_filter = ("location", "category", "status", "company")
    form = EventAdminForm
    search_fields = ("title", "description")
    ordering = ("event_date",)

    def get_fields(self, request, obj=None):
        if obj is None:
            fields = super().get_fields(request, obj)
            fields = [field for field in fields if field != "status"]
            return fields
        return super().get_fields(request, obj)

    def save_model(self, request, obj, form, change):
        if not change:
            default_status = Status.objects.filter(
                title__iexact="Запланировано"
            ).first()
            if default_status:
                obj.status = default_status

        obj.full_clean()
        super().save_model(request, obj, form, change)
